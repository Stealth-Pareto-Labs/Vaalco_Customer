"""
parser.py
=========
The ONLY place a spreadsheet is read.

Reads the REAL midnight-report files as-is — all three sheets:
  - DPR-XXX        : operations hours, full fluid inventory, run-hours,
                     crew movements, HSE tallies, and the hour-by-hour
                     ACTIVITY LOG (free text — the unstructured "why").
  - Technical Info : per-cylinder exhaust temps + deviation, turbocharger,
                     pressures, bearing temps, engine load/RPM.
  - HSE Statistics : the monthly indicator matrix.

Reports vary slightly in row position, so we locate blocks by their labels
rather than fixed cell coordinates. Output is one rich, structured record
per report that preserves both the numbers and the raw text.
"""

import os
import glob
import re
import datetime


def _f(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def _s(x):
    return "" if x is None else str(x).strip()


def _find_row(ws, col, needle, start=1, end=None):
    end = end or ws.max_row
    needle = needle.lower()
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and needle in str(v).lower():
            return r
    return None


def _t(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    return _s(v)


def _parse_dpr(ws):
    out = {}
    out["date_raw"] = _s(ws["H3"].value)
    out["latitude"] = _s(ws["B3"].value)
    out["longitude"] = _s(ws["B4"].value)
    out["wind"] = _s(ws["D3"].value)
    out["sea_state"] = _s(ws["D4"].value)
    out["temperature"] = _f(ws["F3"].value)
    out["current"] = _s(ws["F4"].value)
    out["field"] = _s(ws["H4"].value)
    out["vessel"] = _s(ws["H2"].value)

    ops_hdr = _find_row(ws, 1, "OPERATIONS", 1, 20)
    if ops_hdr:
        vals = [ws.cell(row=ops_hdr + 2, column=c).value for c in range(1, 9)]
        keys = ["SES", "SFS", "AHO", "SOP", "TOP", "IDP", "IAS", "total_hours"]
        out["operations_hours"] = {k: _f(v) for k, v in zip(keys, vals)}

    dp_hdr = _find_row(ws, 1, "HOURS ON DP", 1, 25)
    if dp_hdr:
        vals = [ws.cell(row=dp_hdr + 2, column=c).value for c in range(1, 9)]
        keys = ["SES", "SFS", "AHO", "SOP", "TOP", "IDP", "IAS", "total_dp_hours"]
        out["dp_hours_breakdown"] = {k: _f(v) for k, v in zip(keys, vals)}
        out["dp_hours"] = _f(vals[7])

    fluid_hdr = _find_row(ws, 1, "Fluid type", 1, 40)
    fluids = {}
    if fluid_hdr:
        r = fluid_hdr + 1
        while r <= ws.max_row:
            name = _s(ws.cell(row=r, column=1).value)
            if not name or name.upper().startswith("POB"):
                break
            fluids[name] = {
                "unit": _s(ws.cell(row=r, column=2).value),
                "prev_balance": _f(ws.cell(row=r, column=3).value),
                "consumed": _f(ws.cell(row=r, column=4).value),
                "bunk_in": _f(ws.cell(row=r, column=5).value),
                "bunk_out": _f(ws.cell(row=r, column=6).value),
                "balance": _f(ws.cell(row=r, column=7).value),
                "comment": _s(ws.cell(row=r, column=8).value),
            }
            r += 1
    out["fluids"] = fluids
    if "Fuel - MGO" in fluids:
        out["fuel_m3"] = fluids["Fuel - MGO"].get("consumed")

    pob_hdr = _find_row(ws, 1, "POB  NUMBER", 1, 45)
    if pob_hdr:
        sub = _find_row(ws, 1, "SUBTOTAL", pob_hdr, pob_hdr + 8)
        if sub:
            out["pob_total"] = _f(ws.cell(row=sub, column=5).value)

    mach_hdr = _find_row(ws, 3, "Run hours from last report", 1, ws.max_row)
    machines = {}
    if mach_hdr:
        r = mach_hdr + 2
        while r <= ws.max_row:
            label = _s(ws.cell(row=r, column=1).value)
            if not label or "Personnel" in label:
                break
            low = label.lower()
            if label.startswith("Run hrs") or "compressor" in low or "winch" in low or "crane" in low:
                machines[label] = {
                    "run_hrs_since_last": _f(ws.cell(row=r, column=3).value),
                    "monthly": _f(ws.cell(row=r, column=4).value),
                    "since_overhaul": _f(ws.cell(row=r, column=5).value),
                    "total_run_hours": _f(ws.cell(row=r, column=6).value),
                    "last_lube_change": _s(ws.cell(row=r, column=7).value),
                    "next_lube_change": _s(ws.cell(row=r, column=8).value),
                }
            r += 1
    out["machinery"] = machines

    hse_hdr = _find_row(ws, 1, "Categories", 1, ws.max_row)
    hse = {}
    if hse_hdr:
        r = hse_hdr + 1
        while r <= ws.max_row:
            cat = _s(ws.cell(row=r, column=1).value)
            if not cat or cat.upper().startswith("ACTIVITY"):
                break
            hse[cat] = {
                "to_date": _f(ws.cell(row=r, column=2).value),
                "today": _f(ws.cell(row=r, column=3).value),
                "total": _f(ws.cell(row=r, column=4).value),
            }
            r += 1
    out["hse_tallies"] = hse

    act_hdr = _find_row(ws, 1, "Activity last 24H", 1, ws.max_row)
    activities = []
    if act_hdr:
        r = act_hdr + 2
        while r <= ws.max_row:
            start = ws.cell(row=r, column=1).value
            op = _s(ws.cell(row=r, column=4).value)
            if start is None and not op:
                break
            if op:
                activities.append({
                    "start": _t(start),
                    "finish": _t(ws.cell(row=r, column=2).value),
                    "duration": _t(ws.cell(row=r, column=3).value),
                    "operation": op,
                })
            r += 1
    out["activity_log"] = activities
    out["activity_text"] = " | ".join(a["operation"] for a in activities)
    return out


def _parse_technical(ws):
    out = {}

    def rv(r, c1, c2):
        return [_f(ws.cell(row=r, column=c).value) for c in range(c1, c2 + 1)]

    me1_hdr = _find_row(ws, 2, "Exhaust gas temp", 1, 12)
    if me1_hdr:
        tr = me1_hdr + 2
        out["me1_cyl_temps"] = [v for v in rv(tr, 2, 13) if v is not None]
        out["me1_deviation"] = _f(ws.cell(row=tr, column=14).value)
        out["me2_cyl_temps"] = [v for v in rv(tr, 16, 27) if v is not None]
        out["me2_deviation"] = _f(ws.cell(row=tr, column=28).value)

    out["me1_rpm"] = _f(ws["B5"].value)
    out["me1_load_pct"] = _f(ws["D5"].value)
    out["me2_rpm"] = _f(ws["P5"].value)
    out["me2_load_pct"] = _f(ws["R5"].value)
    out["me1_tc_rpm"] = [_f(ws["B13"].value), _f(ws["H13"].value)]
    out["me1_lo_pressure"] = _f(ws["B17"].value)
    out["me1_fo_pressure"] = _f(ws["H17"].value)
    out["me2_lo_pressure"] = _f(ws["P17"].value)
    out["me2_fo_pressure"] = _f(ws["V17"].value)
    out["dg2_cyl_temps"] = [v for v in rv(26, 16, 27) if v is not None]
    return out


def _parse_hse(ws):
    out = {"indicators": {}}
    hdr = _find_row(ws, 3, "Indicators", 1, 10)
    if not hdr:
        return out
    r = hdr + 1
    blanks = 0
    while r <= ws.max_row and blanks < 5:
        name = _s(ws.cell(row=r, column=3).value)
        if not name:
            blanks += 1
            r += 1
            continue
        blanks = 0
        vals = [_f(ws.cell(row=r, column=c).value) for c in range(5, 40)]
        vals = [v for v in vals if v is not None]
        if vals:
            out["indicators"][name] = {
                "latest": vals[-1],
                "month_total": sum(vals),
            }
        r += 1
    return out


def _iso_date(raw, path):
    raw = str(raw).strip()
    for sep in (".", "/", "-"):
        parts = raw.split(sep)
        if len(parts) == 3 and len(parts[2]) == 4:
            d, m, y = parts
            return f"{y}-{int(m):02d}-{int(d):02d}"
    mt = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", os.path.basename(path))
    if mt:
        d, m, y = mt.groups()
        return f"{y}-{m}-{d}"
    return raw or os.path.basename(path)


def parse_one(path):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("Missing dependency. Run:  pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    record = {"source_file": os.path.basename(path)}

    dpr_sheet = next((s for s in wb.sheetnames if s.upper().startswith("DPR")), wb.sheetnames[0])
    record.update(_parse_dpr(wb[dpr_sheet]))

    if "Technical Info" in wb.sheetnames:
        record["technical"] = _parse_technical(wb["Technical Info"])
    if "HSE Statistics" in wb.sheetnames:
        record["hse_monthly"] = _parse_hse(wb["HSE Statistics"])

    record["date"] = _iso_date(record.get("date_raw", ""), path)
    record["fuel_L"] = round((record.get("fuel_m3") or 0) * 1000)
    return record


def parse_folder(folder):
    records = []
    for path in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        try:
            records.append(parse_one(path))
        except Exception as e:
            print(f"  ! could not parse {os.path.basename(path)}: {e}")
    records.sort(key=lambda d: d.get("date", ""))
    return records
